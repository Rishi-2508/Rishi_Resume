'''
check for  1 excel in firm folders
check for  1 excel in ftp folder
check  xlsx files are delivered 
check date from xlsx on 3/4 sheets
check date in file name for all files

Owner : Rushikesh Shingare 
external Libraries : pandas and openpyxl
'''

import os
import sys
from datetime import datetime,timedelta
from openpyxl import workbook
from openpyxl import load_workbook
import pandas as pd
from pandas.tseries.offsets import BDay

check = ['File_name_1']
result = []
output = []
fdate = datetime.now()
fdateformat = fdate.strftime("%Y%m%d")
t1date = fdate - BDay(1)
sdateformat = t1date.strftime(" %d %B, %Y")
mfilename =[]
xl = [["","QC checks on excel sheet"]]

#set root for required files  and  get created date in proper format
for root,dirs,files in os.walk ("//folder1//folder2//folder3//firm_name//"):
    for each in check:
        for eachfile in files:
            if each in eachfile:
                createdtime = os.path.getctime(root+'//'+eachfile)
                formatedtime = datetime.fromtimestamp(createdtime).strftime("%Y-%m-%d")
                fdatetime = datetime.fromtimestamp(createdtime).strftime("%Y-%m-%d %H:%M:%S")
                today = datetime.now()
                todaystr = today.strftime("%Y-%m-%d")

                #check for date in file name for todays files
                if formatedtime == todaystr and not ("Month" in eachfile):
                        noext = eachfile.split('.')
                        noext.pop
                        mfilename.append(noext[0])
                        for each1 in mfilename:
                            if fdateformat in each1:
                                comment = "Date in file name is valid"
                            else:
                                comment = "Check date in file name"
                        result.append((root,eachfile,fdatetime,comment))

                #open file and check data for todays files
                if formatedtime == todaystr :

                    try:
                        wb = load_workbook(root+"//"+eachfile,'r',data_only = True)
                        for sheet in wb.worksheets:
                            ws = sheet
                            
                            #for sheet named "Report"
                            if sheet.title == 'Report':
                                b2_all = ws['b2'].value
                                b2_split = b2_all.split(" - ")
                                b2 = b2_split[2]
                                s_name = sheet.title
                                if b2 == sdateformat:
                                    scomment = "Date in Sheet is Valid"
                                else:
                                    scomment = "Check Date in Sheet"

                            #for sheet named "Source data"
                            elif sheet.title == 'Source data':
                                b2_s = ws['a2'].value
                                b2=((datetime.utcfromtimestamp(0) + timedelta(b2_s-25569)).strftime(" %d %B, %Y"))
                                s_name = sheet.title
                                if b2 == sdateformat:
                                    scomment = "Date in Sheet is Valid"
                                else:
                                    scomment = "Check Date in Sheet"

                            #for rest of sheets
                            else :    
                                a2 = ws['a2'].value
                                b2 = a2.strftime(" %d %B, %Y")
                                s_name = sheet.title
                                if b2 == sdateformat:
                                    scomment = "Date in Sheet is Valid"
                                else:
                                    scomment = "Check Date in Sheet"
                            output.append((root,s_name,b2,scomment))
                        wb.close()
                    except:
                        pass

#get results in CSV
pandasDF = pd.DataFrame(data = result, columns = ["Path","File_Name","Created_Time","Comment"])

pandasDF1 = pd.DataFrame(data = output, columns = ["Path","Sheet_Name","Date_in_Sheet","Comment"])

pandasDF2 = pd.DataFrame(data = xl, columns = ['',''])

pandasDF.to_csv("//folder1//folder2//folder3//Rishi//folder4//folder5//review_"+fdateformat+".csv")

pandasDF2.to_csv("//folder1//folder2//folder3//Rishi//folder4//folder5//review_"+fdateformat+".csv",mode = 'a',index = False)

pandasDF1.to_csv("//folder1//folder2//folder3//Rishi//folder4//folder5//review_"+fdateformat+".csv",mode = 'a')