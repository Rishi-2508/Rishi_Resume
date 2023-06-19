'''
check csv and xlsx files are delivered 
check fund and date from xlsx file
check file names for xlsxs and csv files

Owner : Rushikesh Shingare (rbs3)
external Libraries : pandas and openpyxl
'''

import os
import sys
from datetime import datetime
from datetime import date
from openpyxl import workbook
from openpyxl import load_workbook
import pandas as pd


#Veriables :
output =[]
result = []
resultopen = []
fdate = datetime.now()
fdateformat = fdate.strftime("%Y%m%d")

#monitoring list:
mlist = ['filename_1','filename_2','filename_3','filename_4','filename_5','filename_6','filename_7','filename_8','filename_9','filename_10']

#walk trough directory
for root, dirs, files in os.walk("//folder1//folder2//reports//Firm_name//FIRM"):
    for mfile in  mlist:
        if mfile in root:
            for eachfile in files:
                createdtime = os.path.getctime(root+"//"+eachfile)
                formateddate = datetime.fromtimestamp(createdtime).strftime("%Y-%m-%d")
                formatedtime = datetime.fromtimestamp(createdtime).strftime("%Y-%m-%d %H:%M:%S")
                today = datetime.now()
                todaystr = today.strftime("%Y-%m-%d")

                #open todays xl file and grab fund name and effective date from it
                if formateddate == todaystr and not ('ME' in eachfile):
                    try:
                        wb = load_workbook(root+"//"+eachfile,'r')
                        ws = wb.active
                        #fund name
                        b7 = ws['b7'].value
                        if str(ws['b7'].value) == '(blank)' or str(ws['b7'].value)=='None':
                            b7 = ws['b9'].value
                            if str(ws['b9'].value)=='(blank)'or str(ws['b9'].value)=='None': 
                                b7 = ws['b10'].value            #Severcopy
                        #print(b7)
                        #as of date    
                        b2all = (ws['b2'].value)
                        b2l = b2all.split('for')
                        b2c = b2l[1]
                        b2 = b2c.replace(",","")
                        wb.close()
                    except:
                        b2 = 'batch_csv'
                        b7 = b2

                #Check Date in file name
                if formateddate == todaystr and not('ME' in eachfile):
                    result.append(eachfile)
                    for each in result:
                        noext = each.split('.')
                        noext.pop()
            
                        for each1 in noext:
                            dateinname = each1.split('_')
                            dateinname.pop(0)
                            dateinname.pop(0)
                            if fdateformat in dateinname:
                                comment = "Date in name is valid"
                            else:
                                comment = "Check date in name!"
                    output.append((eachfile,b7,b2,formatedtime,comment))


pandasDF = pd.DataFrame(data = output,columns = ["File_Name","Fund_In_Report","As of Date","Created_Time(CST)","Comment"])

pandasDF.to_csv('//folder_1//folder_2//folder_3//Rishi//Review//Consol//Macro_Review_'+fdateformat+'.csv')

            